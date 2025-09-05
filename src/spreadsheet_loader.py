import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional
import io

class SpreadsheetLoader:
    """Handles loading spreadsheets from local files only"""
    
    def __init__(self):
        self.supported_formats = ['.xlsx', '.xls', '.csv']
    
    def load_from_file(self, file_path: str) -> Dict[str, Any]:
        """Load spreadsheet from local file"""
        try:
            if file_path.endswith('.csv'):
                return self._load_csv(file_path)
            elif file_path.endswith(('.xlsx', '.xls')):
                return self._load_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format. Supported: {self.supported_formats}")
        except Exception as e:
            raise Exception(f"Error loading file {file_path}: {str(e)}")
    
    def load_from_bytes(self, file_bytes: bytes, filename: str) -> Dict[str, Any]:
        """Load spreadsheet from byte data (for file uploads)"""
        try:
            if filename.endswith('.csv'):
                return self._load_csv_from_bytes(file_bytes)
            elif filename.endswith(('.xlsx', '.xls')):
                return self._load_excel_from_bytes(file_bytes)
            else:
                raise ValueError(f"Unsupported file format: {filename}")
        except Exception as e:
            raise Exception(f"Error loading file {filename}: {str(e)}")
    
    def _load_csv(self, file_path: str) -> Dict[str, Any]:
        """Load CSV file"""
        try:
            df = pd.read_csv(file_path)
            return {
                "name": file_path.split('/')[-1],  # Get filename from path
                "sheets": {
                    "Sheet1": {
                        "data": df.values.tolist(),
                        "headers": df.columns.tolist()
                    }
                }
            }
        except Exception as e:
            raise Exception(f"Error reading CSV file: {str(e)}")
    
    def _load_excel(self, file_path: str) -> Dict[str, Any]:
        """Load Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets = {}
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                data = []
                
                for row in worksheet.iter_rows(values_only=True):
                    # Convert None values to empty strings and handle formulas
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        else:
                            cleaned_row.append(cell)
                    data.append(cleaned_row)
                
                # Remove empty rows at the end
                while data and all(cell == "" or cell is None for cell in data[-1]):
                    data.pop()
                
                # Extract headers from first row if it exists
                headers = data[0] if data else []
                
                sheets[sheet_name] = {
                    "data": data,
                    "headers": headers
                }
            
            return {
                "name": file_path.split('/')[-1],  # Get filename from path
                "sheets": sheets
            }
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def _load_csv_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """Load CSV from bytes"""
        try:
            df = pd.read_csv(io.BytesIO(file_bytes))
            return {
                "name": "uploaded_csv",
                "sheets": {
                    "Sheet1": {
                        "data": df.values.tolist(),
                        "headers": df.columns.tolist()
                    }
                }
            }
        except Exception as e:
            raise Exception(f"Error reading CSV from bytes: {str(e)}")
    
    def _load_excel_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """Load Excel file from bytes"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets = {}
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                data = []
                
                for row in worksheet.iter_rows(values_only=True):
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        else:
                            cleaned_row.append(cell)
                    data.append(cleaned_row)
                
                # Remove empty rows at the end
                while data and all(cell == "" or cell is None for cell in data[-1]):
                    data.pop()
                
                # Extract headers from first row if it exists
                headers = data[0] if data else []
                
                sheets[sheet_name] = {
                    "data": data,
                    "headers": headers
                }
            
            return {
                "name": "uploaded_excel",
                "sheets": sheets
            }
        except Exception as e:
            raise Exception(f"Error reading Excel from bytes: {str(e)}")
    
    def validate_spreadsheet_data(self, data: Dict[str, Any]) -> bool:
        """Validate that spreadsheet data is properly formatted"""
        try:
            if not isinstance(data, dict):
                return False
            
            if "sheets" not in data:
                return False
            
            sheets = data["sheets"]
            if not isinstance(sheets, dict):
                return False
            
            for sheet_name, sheet_data in sheets.items():
                if not isinstance(sheet_data, dict):
                    return False
                
                if "data" not in sheet_data:
                    return False
                
                if not isinstance(sheet_data["data"], list):
                    return False
            
            return True
            
        except Exception:
            return False
    
    def get_spreadsheet_summary(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Get summary information about the loaded spreadsheet"""
        try:
            total_cells = 0
            total_rows = 0
            sheet_names = []
            
            for sheet_name, sheet_data in data["sheets"].items():
                sheet_names.append(sheet_name)
                if "data" in sheet_data:
                    sheet_rows = len(sheet_data["data"])
                    total_rows += sheet_rows
                    for row in sheet_data["data"]:
                        total_cells += len([cell for cell in row if cell != "" and cell is not None])
            
            return {
                "name": data.get("name", "Unknown"),
                "total_sheets": len(data["sheets"]),
                "sheet_names": sheet_names,
                "total_rows": total_rows,
                "total_cells": total_cells
            }
            
        except Exception as e:
            return {
                "name": "Unknown",
                "total_sheets": 0,
                "sheet_names": [],
                "total_rows": 0,
                "total_cells": 0,
                "error": str(e)
            }
    
    def get_sheet_data(self, data: Dict[str, Any], sheet_name: str) -> Optional[Dict[str, Any]]:
        """Get data for a specific sheet"""
        try:
            if sheet_name in data["sheets"]:
                return data["sheets"][sheet_name]
            return None
        except Exception:
            return None
    
    def get_all_sheet_names(self, data: Dict[str, Any]) -> List[str]:
        """Get list of all sheet names"""
        try:
            return list(data["sheets"].keys())
        except Exception:
            return []
    
    def preview_data(self, data: Dict[str, Any], sheet_name: str = None, max_rows: int = 5) -> Dict[str, Any]:
        """Get a preview of the spreadsheet data"""
        try:
            if sheet_name:
                # Preview specific sheet
                if sheet_name not in data["sheets"]:
                    return {"error": f"Sheet '{sheet_name}' not found"}
                
                sheet_data = data["sheets"][sheet_name]["data"]
                preview_data = sheet_data[:max_rows] if len(sheet_data) > max_rows else sheet_data
                
                return {
                    "sheet_name": sheet_name,
                    "preview_data": preview_data,
                    "total_rows": len(sheet_data),
                    "showing_rows": len(preview_data)
                }
            else:
                # Preview first sheet
                first_sheet_name = list(data["sheets"].keys())[0]
                return self.preview_data(data, first_sheet_name, max_rows)
                
        except Exception as e:
            return {"error": str(e)}